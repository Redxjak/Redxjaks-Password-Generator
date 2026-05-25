#!/usr/bin/env python
# Created by Redxjak

# Password Generator


import argparse
import csv
import json
import os
import random
import string
import sys
from datetime import datetime
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import Workbook, load_workbook


VERSION = "1.1.2"
SYMBOLS = "!@#$%^&*()"
EXPORT_EXTENSIONS = {
    "excel": ".xlsx",
    "csv": ".csv",
    "txt": ".txt",
    "json": ".json",
    "markdown": ".md",
}


def remove_excluded_characters(characters, excluded_characters):
    return "".join(character for character in characters if character not in excluded_characters)


def generate_password(min_length, max_length, min_uppercase, min_lowercase, min_symbols, excluded_characters=""):
    uppercase = remove_excluded_characters(string.ascii_uppercase, excluded_characters)
    lowercase = remove_excluded_characters(string.ascii_lowercase, excluded_characters)
    numbers = remove_excluded_characters(string.digits, excluded_characters)
    symbols = remove_excluded_characters(SYMBOLS, excluded_characters)

    if min_uppercase > 0 and uppercase == "":
        raise ValueError("Minimum uppercase letters were requested, but all uppercase letters are excluded.")

    if min_lowercase > 0 and lowercase == "":
        raise ValueError("Minimum lowercase letters were requested, but all lowercase letters are excluded.")

    if min_symbols > 0 and symbols == "":
        raise ValueError("Minimum special characters were requested, but all special characters are excluded.")

    characters = uppercase + lowercase + numbers + symbols

    if characters == "":
        raise ValueError("No characters are available after applying exclusions.")

    password_characters = []
    length = random.randint(min_length, max_length)

    for i in range(min_uppercase):
        password_characters.append(random.choice(uppercase))

    for i in range(min_lowercase):
        password_characters.append(random.choice(lowercase))

    for i in range(min_symbols):
        password_characters.append(random.choice(symbols))

    minimum_required_length = len(password_characters)

    if minimum_required_length > length:
        length = minimum_required_length

    remaining_length = length - len(password_characters)

    for i in range(remaining_length):
        password_characters.append(random.choice(characters))

    random.shuffle(password_characters)

    return "".join(password_characters)


def generate_password_records(
    site_name,
    username,
    count,
    min_length,
    max_length,
    min_uppercase,
    min_lowercase,
    min_symbols,
    excluded_characters=""
):
    records = []

    for i in range(count):
        records.append({
            "Website/Application": site_name,
            "Username": username,
            "Password": generate_password(
                min_length,
                max_length,
                min_uppercase,
                min_lowercase,
                min_symbols,
                excluded_characters
            ),
            "Timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        })

    return records


def get_export_format_from_path(file_path):
    extension = os.path.splitext(file_path)[1].lower()

    for export_format, expected_extension in EXPORT_EXTENSIONS.items():
        if extension == expected_extension:
            return export_format

    return "excel"


def save_records(file_path, records, export_format):
    if export_format == "excel":
        save_excel_records(file_path, records)
    elif export_format == "csv":
        save_csv_records(file_path, records)
    elif export_format == "txt":
        save_text_records(file_path, records)
    elif export_format == "json":
        save_json_records(file_path, records)
    elif export_format == "markdown":
        save_markdown_records(file_path, records)
    else:
        raise ValueError("Unsupported export format.")


def save_excel_records(file_path, records):
    if os.path.exists(file_path):
        workbook = load_workbook(file_path)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Passwords"
        sheet.append(["Website/Application", "Username", "Password", "Timestamp"])

    for record in records:
        sheet.append([
            record["Website/Application"],
            record["Username"],
            record["Password"],
            record["Timestamp"],
        ])

    workbook.save(file_path)


def save_csv_records(file_path, records):
    file_exists = os.path.exists(file_path)

    with open(file_path, "a", newline="", encoding="utf-8") as file:
        writer = csv.DictWriter(
            file,
            fieldnames=["Website/Application", "Username", "Password", "Timestamp"]
        )

        if not file_exists:
            writer.writeheader()

        writer.writerows(records)


def save_text_records(file_path, records):
    with open(file_path, "a", encoding="utf-8") as file:
        for record in records:
            file.write("=====================================\n")
            file.write(f"Website/Application: {record['Website/Application']}\n")
            file.write(f"Username: {record['Username']}\n")
            file.write(f"Password: {record['Password']}\n")
            file.write(f"Timestamp: {record['Timestamp']}\n")


def save_json_records(file_path, records):
    existing_records = []

    if os.path.exists(file_path):
        with open(file_path, "r", encoding="utf-8") as file:
            try:
                existing_records = json.load(file)
            except json.JSONDecodeError:
                existing_records = []

    existing_records.extend(records)

    with open(file_path, "w", encoding="utf-8") as file:
        json.dump(existing_records, file, indent=4)


def save_markdown_records(file_path, records):
    file_exists = os.path.exists(file_path)

    with open(file_path, "a", encoding="utf-8") as file:
        if not file_exists:
            file.write("| Website/Application | Username | Password | Timestamp |\n")
            file.write("| --- | --- | --- | --- |\n")

        for record in records:
            file.write(
                f"| {record['Website/Application']} | {record['Username']} | "
                f"{record['Password']} | {record['Timestamp']} |\n"
            )


def copy_text_to_clipboard(text):
    clipboard_window = tk.Tk()
    clipboard_window.withdraw()
    clipboard_window.clipboard_clear()
    clipboard_window.clipboard_append(text)
    clipboard_window.update()
    clipboard_window.destroy()


def show_error(title, message):
    try:
        if "window" in globals() and window.winfo_exists():
            messagebox.showerror(title, message, parent=window)
            return
    except tk.TclError:
        pass

    print(f"{title}: {message}", file=sys.stderr)


def show_info(title, message):
    try:
        if "window" in globals() and window.winfo_exists():
            messagebox.showinfo(title, message, parent=window)
            return
    except tk.TclError:
        pass

    print(f"{title}: {message}")


def is_number_or_blank(value):
    return value == "" or value.isdigit()


def choose_file_location():
    export_format = export_format_var.get()
    default_extension = EXPORT_EXTENSIONS[export_format]
    filetypes = [
        ("Selected Format", f"*{default_extension}"),
        ("All Files", "*.*"),
    ]

    file_path = filedialog.asksaveasfilename(
        title="Choose Save Location",
        defaultextension=default_extension,
        filetypes=filetypes
    )

    if file_path:
        save_location_entry.delete(0, tk.END)
        save_location_entry.insert(0, file_path)


def create_password():
    site_name = site_entry.get()
    username = username_entry.get()
    save_location = save_location_entry.get()
    excluded_characters = excluded_characters_entry.get()
    export_format = export_format_var.get()

    try:
        min_length = int(min_length_entry.get() or 10)
        max_length = int(max_length_entry.get() or 20)
        min_uppercase = int(min_uppercase_entry.get() or 0)
        min_lowercase = int(min_lowercase_entry.get() or 0)
        min_symbols = int(min_symbols_entry.get() or 0)
        count = int(count_entry.get() or 1)
    except ValueError:
        show_error(
            "Error",
            "Length, count, and minimum requirements must be numbers."
        )
        return

    if min_length < 1 or max_length < 1:
        show_error("Error", "Password length must be at least 1.")
        return

    if min_uppercase < 0 or min_lowercase < 0 or min_symbols < 0:
        show_error("Error", "Minimum requirements cannot be negative.")
        return

    if min_length > max_length:
        show_error(
            "Error",
            "Minimum length cannot be greater than maximum length."
        )
        return

    if count < 1:
        show_error("Error", "Number of passwords must be at least 1.")
        return

    if site_name == "" or username == "" or save_location == "":
        show_error("Error", "Please fill out all fields.")
        return

    try:
        records = generate_password_records(
            site_name,
            username,
            count,
            min_length,
            max_length,
            min_uppercase,
            min_lowercase,
            min_symbols,
            excluded_characters
        )
    except ValueError as error:
        show_error("Error", str(error))
        return

    passwords = "\n".join(record["Password"] for record in records)
    password_output.delete("1.0", tk.END)
    password_output.insert("1.0", passwords)

    try:
        save_records(save_location, records, export_format)
    except PermissionError:
        show_error(
            "File is open",
            "Please close the file before generating a password."
        )
        return

    show_info("Password Generated", "Password Generated Successfully!")


def copy_generated_passwords():
    passwords = password_output.get("1.0", tk.END).strip()

    if passwords == "":
        show_error("Error", "There are no generated passwords to copy.")
        return

    window.clipboard_clear()
    window.clipboard_append(passwords)
    window.update()
    show_info("Copied", "Generated password text copied to the clipboard.")


def run_cli():
    parser = argparse.ArgumentParser(description="Redxjak's Password Generator")
    parser.add_argument(
        "-v",
        "--version",
        action="version",
        version=f"Redxjak's Password Generator version {VERSION}"
    )
    parser.add_argument("--site", default="Command Line", help="Website or application name.")
    parser.add_argument("--username", default="CLI User", help="Username for the generated record.")
    parser.add_argument("--min-length", type=int, default=10, help="Minimum password length.")
    parser.add_argument("--max-length", type=int, default=20, help="Maximum password length.")
    parser.add_argument("--min-uppercase", type=int, default=0, help="Minimum uppercase letters.")
    parser.add_argument("--min-lowercase", type=int, default=0, help="Minimum lowercase letters.")
    parser.add_argument("--min-symbols", type=int, default=0, help="Minimum special characters.")
    parser.add_argument("--exclude", default="", help="Characters to exclude, such as O0l1.")
    parser.add_argument("--count", type=int, default=1, help="Number of passwords to generate.")
    parser.add_argument("--output", help="Optional output file path.")
    parser.add_argument(
        "--format",
        choices=["excel", "csv", "txt", "json", "markdown"],
        help="Output format. Defaults to the output file extension."
    )
    parser.add_argument("--copy", action="store_true", help="Copy generated passwords to clipboard.")
    args = parser.parse_args()

    if args.min_length > args.max_length:
        parser.error("Minimum length cannot be greater than maximum length.")

    if args.min_length < 1 or args.max_length < 1:
        parser.error("Password length must be at least 1.")

    if args.min_uppercase < 0 or args.min_lowercase < 0 or args.min_symbols < 0:
        parser.error("Minimum requirements cannot be negative.")

    if args.count < 1:
        parser.error("Count must be at least 1.")

    try:
        records = generate_password_records(
            args.site,
            args.username,
            args.count,
            args.min_length,
            args.max_length,
            args.min_uppercase,
            args.min_lowercase,
            args.min_symbols,
            args.exclude
        )
    except ValueError as error:
        parser.error(str(error))
    passwords = "\n".join(record["Password"] for record in records)

    print(passwords)

    if args.output:
        export_format = args.format or get_export_format_from_path(args.output)
        save_records(args.output, records, export_format)

    if args.copy:
        copy_text_to_clipboard(passwords)


def run_gui():
    global window
    global site_entry
    global username_entry
    global min_length_entry
    global max_length_entry
    global min_uppercase_entry
    global min_lowercase_entry
    global min_symbols_entry
    global excluded_characters_entry
    global count_entry
    global export_format_var
    global save_location_entry
    global password_output

    window = tk.Tk()
    window.title("Redxjak's Password Generator")
    window.geometry("430x720")
    number_validator = (window.register(is_number_or_blank), "%P")

    tk.Label(window, text="Redxjak's Password Generator", font=("Arial", 16)).pack(pady=10)

    tk.Label(window, text="Website/Application Name:").pack()
    site_entry = tk.Entry(window, width=45)
    site_entry.pack()

    tk.Label(window, text="Username:").pack()
    username_entry = tk.Entry(window, width=45)
    username_entry.pack()

    tk.Label(window, text="Minimum Password Length: (Default 10)").pack()
    min_length_entry = tk.Entry(window, width=45, validate="key", validatecommand=number_validator)
    min_length_entry.pack()

    tk.Label(window, text="Maximum Password Length: (Default 20)").pack()
    max_length_entry = tk.Entry(window, width=45, validate="key", validatecommand=number_validator)
    max_length_entry.pack()

    tk.Label(window, text="Minimum Uppercase Letters:").pack()
    min_uppercase_entry = tk.Entry(window, width=45, validate="key", validatecommand=number_validator)
    min_uppercase_entry.pack()

    tk.Label(window, text="Minimum Lowercase Letters:").pack()
    min_lowercase_entry = tk.Entry(window, width=45, validate="key", validatecommand=number_validator)
    min_lowercase_entry.pack()

    tk.Label(window, text="Minimum Special Characters:").pack()
    min_symbols_entry = tk.Entry(window, width=45, validate="key", validatecommand=number_validator)
    min_symbols_entry.pack()

    tk.Label(window, text="Characters to Exclude:").pack()
    excluded_characters_entry = tk.Entry(window, width=45)
    excluded_characters_entry.insert(0, "O0l1")
    excluded_characters_entry.pack()

    tk.Label(window, text="Number of Passwords: (Default 1)").pack()
    count_entry = tk.Entry(window, width=45, validate="key", validatecommand=number_validator)
    count_entry.pack()

    tk.Label(window, text="Export Format:").pack()
    export_format_var = tk.StringVar(value="excel")
    tk.OptionMenu(window, export_format_var, "excel", "csv", "txt", "json", "markdown").pack()

    tk.Label(window, text="Save Location:").pack()

    tk.Label(
        window,
        text="*Click 'Choose Save Location' and enter a file name, or select a previous file to add to it.*",
        wraplength=350,
        justify="center",
        fg="gray",
        font=("Arial", 8)
    ).pack()

    save_location_entry = tk.Entry(window, width=45)
    save_location_entry.pack()

    browse_button = tk.Button(
        window,
        text="Choose Save Location",
        command=choose_file_location
    )
    browse_button.pack(pady=5)

    generate_button = tk.Button(
        window,
        text="Generate Password",
        command=create_password
    )
    generate_button.pack(pady=5)

    copy_button = tk.Button(
        window,
        text="Copy Generated Passwords",
        command=copy_generated_passwords
    )
    copy_button.pack(pady=5)

    tk.Label(window, text="Generated Password:").pack()
    password_output = tk.Text(window, width=45, height=4)
    password_output.pack(pady=5)

    window.mainloop()


def main():
    if len(sys.argv) > 1:
        run_cli()
    else:
        try:
            run_gui()
        except tk.TclError as error:
            print(
                f"Redxjak's Password Generator version {VERSION}\n"
                "This program should be run with a GUI.\n"
                "For command-line usage, run with --help to see available options.",
                file=sys.stderr
            )
            print(f"GUI error: {error}", file=sys.stderr)
            sys.exit(1)


if __name__ == "__main__":
    main()
